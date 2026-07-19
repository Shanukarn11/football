from __future__ import annotations
from django.contrib import admin
from sre_parse import State
from django.contrib import admin
from django.contrib import messages
import csv
from django.http import HttpResponse
from .models import HomePage, SupportedByImages2526, TeamVideoCarousel,Resources,WorkshopInternationalLogo,WorkshopNews, About_TeamAdvisors, Calendar_Season3, Gallery_Buttons, Photo_Gallery_Images,InstaFeed,Previous_Workshop_Images,Clubs_Parter_abouttrials, UserFlow_Season3, Winners_Season2,Winners_Workshop_Images,Testimonials, Season3_Features,Associate_TournamentOrganizer_Right,TakeaWay1_TournamentOrganizer,TakeaWay2_TournamentOrganizer,Associate_TournamentOrganizer_Images, TabelSeason3, Clubs_Parter_Season2, Season3_Images, About_TeamCore_Team, Selected_Players_Season2, Clubs_Parter_Season1, Clubs_Parter_Home, About_OURTEAM_of_IkF, TakeaWay1_Clubs, TakeaWay2_Clubs, TakeaWay3_Clubs, Associate_Clubs_Images, Associate_Clubs_Right, TakeaWay1_Academy, Associate_Academy_Images, TakeaWay2_Academy, About_Us_Images, About_section1, AboutTrials_BannerPhoto, AboutTrials_Feature, AboutTrials_Images, AboutUs, Associate_Corporate_Images, Associate_Corporate_Main, Associate_Corporate_Right,  Associate_PSUS_Images, Associate_PSUS_Main, Associate_PSUS_Right, Associate_Philanthropists_Images, Associate_Philanthropists_Main, Associate_Philanthropists_Right, Buttons_IKF_FOR_ALL, Clubs, Clubs_Partners1, Corporate_Images, Corporate_Partners1, IKF_Tech1, IKF_Tech2, IKF_Tech3, IKF_Tech4, IKF_Technology_Images, Initiatives, International_Images, International_Partners1, MasterLabels, National_Images, National_Partners1, NavTabs, Navbar, Lang, HomeBanner, HomeImages, News, Partners, Photos_IKF_FOR_ALL, PlayAbroad_Images, Playabroad, Playabroad_Features, Playabroad_Initiatives, Reginoal_Images, Regional_Partners1, Scouting_Features, Scouting_Initiatives, ScoutingCertification_Images, Season1_Bottom, Season1_Button, Season1_Feature, Season1_Images, Season1_Tabs, Season2_Features, Season2_Images, Selected_Players_Season1, SubTab, Tab, TabelPlayabroad, TabelScoutingCertification, TabelSeason2, TabelWorkshops, TabelWorkshopsReg, TakeaWay1_Corporate, TakeaWay1_PSUS, TakeaWay1_Philanthropists, TakeaWay2_Corporate, TakeaWay2_PSUS, TakeaWay2_Philanthropists, TakeaWay3_PSUS, UpCommingMatchs, Winners, WorkShops_Images, WorkShopsButton, WorkShopsExperts, WorkShopsFeatureStrip, MasterSeason, MasterState, MasterCity, MasterPosition, MasterIkfSeasonUi, TrialsAndInitiativeType, TrialsAndInitiativeNav, ActiveStatusNav, WorkShopsReg_Images, WorkShopsRegButton, WorkShopsRegExperts, WorkShopsRegFeatureStrip, Youtube,TabelPastTrial, SupportedBy, SupportedByFinancial, SupportedByImages, MoreaboutIKF ,Donate_Images,DonateFeatureStrip,Donation,TabelDonate,PaymentFeatureStrip, SupportIkfFeatureStrip, PartnerSeasonSupportIKF,Calender, Finalists,Mentor, Coaches, Scout

from django.contrib import admin
from django.utils.html import format_html

from .models import (
    RepTemplate, RepHighlight, RepPlayer, RepSocialLink, MasterLabels
)
from django.utils.html import format_html
from django.urls import reverse
from django.conf import settings

# main/admin.py
from django.contrib import admin, messages
from django.utils import timezone
from django.utils.html import format_html
from django.urls import reverse, NoReverseMatch
from django.conf import settings

from .models import RepTemplate, RepHighlight, RepPlayer, RepSocialLink, MasterLabels
# ikf/main/admin.py
from django.contrib import admin, messages
from django import forms
from django.utils import timezone
from django.utils.html import format_html
from django.urls import reverse, NoReverseMatch
from django.conf import settings


# main/admin.py
from django import forms
from django.contrib import admin
from django.contrib.admin.widgets import FilteredSelectMultiple
from django.utils.html import format_html
from django.urls import reverse, NoReverseMatch
from django.utils import timezone
from django.conf import settings

from .models import (
    RepTemplate, RepHighlight, RepPlayer, RepSocialLink,
    Status, ONBOARDING_FIELDS, FIELD_SLUGS
)

# ---------- Group fields into sections (purely for display) ----------
SECTIONS = {
    "Branding": ["rep_logo", "rep_banner", "players_bg"],
    "Theme": ["theme", "show_pills", "show_stats", "brand_color", "hero_bg_color", "card_color", "text_color"],
    "Stats": ["stat1_value","stat1_label","stat2_value","stat2_label","stat3_value","stat3_label","stat4_value","stat4_label"],
    "Meta": ["cities_of_operation","website_url","rep_founder"],
    "Narratives": ["content","rep_opinion","rep_category","rep_category_brief","csr_activities","ikf_activities"],
}
LABEL_BY_SLUG = dict(ONBOARDING_FIELDS)

def grouped_choices():
    """Produce optgroup choices for a nicer FilteredSelectMultiple."""
    groups = []
    for title, slugs in SECTIONS.items():
        opts = [(s, LABEL_BY_SLUG.get(s, s)) for s in slugs if s in FIELD_SLUGS]
        if opts:
            groups.append((title, opts))
    return groups


# main/admin.py
from django import forms
from django.contrib import admin
from django.utils import timezone
from django.utils.html import format_html
from django.urls import reverse, NoReverseMatch
from django.conf import settings

from .models import (
    RepTemplate, RepHighlight, RepPlayer, RepSocialLink,
    Status, ONBOARDING_FIELDS, FIELD_SLUGS
)

# ------------- Grouping for nicer layout -------------
# ===========================
# Status (Plan) admin (NO JSON)
# ===========================

# Group fields into sections for nicer checkbox groups
SECTIONS = {
    "Branding":   ["rep_logo", "rep_banner", "players_bg"],
    "Theme":      ["theme", "show_pills", "show_stats", "brand_color", "hero_bg_color", "card_color", "text_color"],
    "Stats":      ["stat1_value","stat1_label","stat2_value","stat2_label","stat3_value","stat3_label","stat4_value","stat4_label"],
    "Meta":       ["cities_of_operation","website_url","rep_founder"],
    "Narratives": ["content","rep_opinion","rep_category","rep_category_brief","csr_activities","ikf_activities"],
}
LABEL_BY_SLUG = dict(ONBOARDING_FIELDS)

CHOICES_BRANDING   = [(s, LABEL_BY_SLUG.get(s, s)) for s in SECTIONS["Branding"]]
CHOICES_THEME      = [(s, LABEL_BY_SLUG.get(s, s)) for s in SECTIONS["Theme"]]
CHOICES_STATS      = [(s, LABEL_BY_SLUG.get(s, s)) for s in SECTIONS["Stats"]]
CHOICES_META       = [(s, LABEL_BY_SLUG.get(s, s)) for s in SECTIONS["Meta"]]
CHOICES_NARRATIVES = [(s, LABEL_BY_SLUG.get(s, s)) for s in SECTIONS["Narratives"]]

class StatusAdminForm(forms.ModelForm):
    # Basics
    name = forms.SlugField(
        label="Name (lowercase only)",
        help_text="Lowercase key used to match RepTemplate ‘next_status’ (e.g. premium, normal).",
        max_length=32,
    )
    label = forms.CharField(
        label="Label",
        help_text="Friendly name visible to reps/admins (e.g. Premium).",
        max_length=64,
        required=False,
    )
    # Checkbox groups
    enabled_branding = forms.MultipleChoiceField(
        label="Branding", required=False, choices=CHOICES_BRANDING,
        widget=forms.CheckboxSelectMultiple(attrs={"class": "checkbox-grid"}),
    )
    enabled_theme = forms.MultipleChoiceField(
        label="Theme", required=False, choices=CHOICES_THEME,
        widget=forms.CheckboxSelectMultiple(attrs={"class": "checkbox-grid"}),
    )
    enabled_stats = forms.MultipleChoiceField(
        label="Stats", required=False, choices=CHOICES_STATS,
        widget=forms.CheckboxSelectMultiple(attrs={"class": "checkbox-grid"}),
    )
    enabled_meta = forms.MultipleChoiceField(
        label="Meta", required=False, choices=CHOICES_META,
        widget=forms.CheckboxSelectMultiple(attrs={"class": "checkbox-grid"}),
    )
    enabled_narratives = forms.MultipleChoiceField(
        label="Narratives", required=False, choices=CHOICES_NARRATIVES,
        widget=forms.CheckboxSelectMultiple(attrs={"class": "checkbox-grid"}),
    )
    # Optional explicit ordering (subset of enabled)
    field_order_csv = forms.CharField(
        label="Field order (CSV)",
        required=False,
        help_text="Optional: comma-separated field names to control order (must be a subset of Enabled).",
    )

    class Meta:
        model = Status
        fields = ("name", "label")

    class Media:
        css = {"all": ("main/admin_status.css",)}  # optional

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        enabled = list(self.instance.enabled_fields or [])
        self.fields["enabled_branding"].initial   = [s for s in SECTIONS["Branding"]   if s in enabled]
        self.fields["enabled_theme"].initial      = [s for s in SECTIONS["Theme"]      if s in enabled]
        self.fields["enabled_stats"].initial      = [s for s in SECTIONS["Stats"]      if s in enabled]
        self.fields["enabled_meta"].initial       = [s for s in SECTIONS["Meta"]       if s in enabled]
        self.fields["enabled_narratives"].initial = [s for s in SECTIONS["Narratives"] if s in enabled]
        if enabled:
            self.fields["field_order_csv"].initial = ",".join(enabled)

    def clean(self):
        cleaned = super().clean()

        # Collect enabled from groups, keep order & dedupe
        seq = (
            cleaned.get("enabled_branding", []) +
            cleaned.get("enabled_theme", []) +
            cleaned.get("enabled_stats", []) +
            cleaned.get("enabled_meta", []) +
            cleaned.get("enabled_narratives", [])
        )
        seen, enabled_flat = set(), []
        for s in seq:
            if s not in seen:
                enabled_flat.append(s)
                seen.add(s)

        # Validate names are known
        unknown = set(enabled_flat) - set(FIELD_SLUGS)
        if unknown:
            raise forms.ValidationError(f"Unknown fields: {', '.join(sorted(unknown))}")

        # Optional order CSV — if provided, reorder enabled fields to match it
        order_csv = (cleaned.get("field_order_csv") or "").strip()
        if order_csv:
            requested = [x.strip() for x in order_csv.split(",") if x.strip()]
            # keep only enabled and dedupe
            requested = [x for x in requested if x in enabled_flat]
            # append any enabled that weren't in requested
            enabled_ordered = requested + [x for x in enabled_flat if x not in requested]
        else:
            enabled_ordered = enabled_flat

        # Save back to model (this writes pipe-CSV under the hood)
        self.instance.enabled_fields = enabled_ordered

        # Ensure slug normalized
        if self.instance.name:
            self.instance.name = self.instance.name.lower()

        return cleaned

@admin.register(Status)
class StatusAdmin(admin.ModelAdmin):
    form = StatusAdminForm
    list_display = ("name", "label", "enabled_count")

    fieldsets = (
        ("Basics", {"fields": (("name", "label"),)}),
        ("Onboarding fields", {
            "fields": (
                "enabled_branding",
                "enabled_theme",
                "enabled_stats",
                "enabled_meta",
                "enabled_narratives",
            ),
            "description": "Pick which fields appear on the onboarding form.",
            "classes": ("collapse",),
        }),
        ("Field order", {"fields": ("field_order_csv",)}),
    )

    @admin.display(description="Enabled")
    def enabled_count(self, obj):
        return len(obj.enabled_fields or [])


# ===========================
# RepTemplate admin (with inlines)
# ===========================

class RepHighlightInline(admin.TabularInline):
    model = RepHighlight
    extra = 0
    fields = ("order", "title", "text")
    ordering = ("order",)

class RepPlayerInline(admin.TabularInline):
    model = RepPlayer
    extra = 0
    fields = ("order", "name", "subtitle", "photo", "bio")
    ordering = ("order",)

class RepSocialLinkInline(admin.TabularInline):
    model = RepSocialLink
    extra = 0
    fields = ("platform", "url")

@admin.action(description="Export onboarding links (CSV)")
def export_onboarding_links(modeladmin, request, queryset):
    base = getattr(settings, "SITE_BASE_URL", request.build_absolute_uri("/").rstrip("/"))
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="rep_onboarding_links.csv"'
    w = csv.writer(resp)
    w.writerow(["rep_name", "rep_num", "onboarding_token", "onboarding_url"])
    for rep in queryset:
        path = reverse("rep_onboarding", args=[rep.onboarding_token])
        w.writerow([rep.rep_name, rep.rep_num, rep.onboarding_token, f"{base}{path}"])
    return resp

@admin.register(RepTemplate)
class RepTemplateAdmin(admin.ModelAdmin):
    list_display = (
        "rep_name", "rep_num", "next_status",
        "is_approved", "is_submitted", "created_at",
        "onboarding_link", "preview_link",
    )
    list_filter  = ("next_status", "is_approved", "is_submitted", "created_at")
    search_fields = ("rep_name", "rep_num", "template_link", "onboarding_token")
    readonly_fields = (
        "onboarding_token", "created_at", "submitted_at",
        "approved_at", "approved_by", "onboarding_link", "preview_link", "public_link",
    )
    actions = [export_onboarding_links]
    inlines = [RepHighlightInline, RepPlayerInline, RepSocialLinkInline]

    fieldsets = (
        ("Identity", {"fields": (("rep_name", "rep_num"), "template_link")}),
        ("Visuals", {"fields": ("rep_logo", "rep_banner", "players_bg")}),
        ("Theme & Colors", {"fields": ("theme", ("brand_color","hero_bg_color","card_color","text_color"), ("show_pills","show_stats"))}),
        ("Stats", {"fields": (("stat1_value","stat1_label"), ("stat2_value","stat2_label"),
                              ("stat3_value","stat3_label"), ("stat4_value","stat4_label"))}),
        ("Meta", {"fields": ("cities_of_operation","website_url","rep_founder","content","rep_opinion","rep_category","rep_category_brief","csr_activities","ikf_activities")}),
        ("Workflow", {"fields": (("next_status","is_locked"),
                                 ("is_submitted","submitted_at"),
                                 ("is_approved","approved_at","approved_by"))}),
        ("System", {"fields": (("onboarding_token", "created_at"), "onboarding_link", "preview_link", "public_link")}),
    )

    # Handy links
    @admin.display(description="Onboarding URL")
    def onboarding_link(self, obj):
        try:
            url = reverse("rep_onboarding", args=[obj.onboarding_token])
            base = getattr(settings, "SITE_BASE_URL", "")
            full = f"{base}{url}" if base else url
            return format_html('<a href="{}" target="_blank" rel="noopener">Open</a>', full)
        except NoReverseMatch:
            return "-"

    @admin.display(description="Preview URL")
    def preview_link(self, obj):
        try:
            url = reverse("rep_preview", args=[obj.onboarding_token])
            base = getattr(settings, "SITE_BASE_URL", "")
            full = f"{base}{url}" if base else url
            return format_html('<a href="{}" target="_blank" rel="noopener">Open</a>', full)
        except NoReverseMatch:
            return "-"

    @admin.display(description="Public URL")
    def public_link(self, obj):
        try:
            url = reverse("rep_public", args=[obj.template_link])
            base = getattr(settings, "SITE_BASE_URL", "")
            full = f"{base}{url}" if base else url
            return format_html('<a href="{}" target="_blank" rel="noopener">Open</a>', full)
        except NoReverseMatch:
            return "-"

from django.contrib import admin
from django.utils.html import format_html
from .models import Lawyer, ScoutProfile

@admin.register(Lawyer)
class LawyerAdmin(admin.ModelAdmin):
    list_display = ("thumb", "name", "city", "expertise_area", "is_featured", "template_link")
    list_display_links = ("thumb", "name")
    search_fields = ("name", "city", "expertise_area", "template_link")
    list_filter = ("city", "is_featured")
    prepopulated_fields = {"template_link": ("name",)}
    readonly_fields = ("preview",)
    fields = ("name", "template_link", "photo", "preview", "city", "expertise_area", "linkedin_url", "is_featured")

    def thumb(self, obj):
        if obj.photo:
            return format_html('<img src="{}" style="height:38px;width:38px;object-fit:cover;border-radius:6px;" />', obj.photo.url)
        return "—"

    def preview(self, obj):
        if obj.photo:
            return format_html('<img src="{}" style="max-width:220px;border-radius:10px;" />', obj.photo.url)
        return "No image"


@admin.register(ScoutProfile)
class ScoutProfileAdmin(admin.ModelAdmin):
    list_display = ("thumb", "name", "city", "state", "experience", "certification", "is_featured", "is_active", "sort_order")
    list_display_links = ("thumb", "name")
    list_editable = ("is_featured", "is_active", "sort_order")
    search_fields = ("name", "city", "state", "regions", "strengths", "certification", "email", "phone")
    list_filter = ("state", "is_featured", "is_active")
    prepopulated_fields = {"slug": ("name",)}
    readonly_fields = ("preview",)
    fieldsets = (
        ("Profile Controls", {
            "fields": ("name", "slug", "photo", "preview", "is_featured", "is_active", "sort_order")
        }),
        ("Personal Information", {
            "fields": ("city", "state", "age_date_of_birth", "languages", "willingness_to_travel", "phone", "email")
        }),
        ("IKF Certification Credentials", {
            "fields": ("certification", "level1_completion_date", "level2_completion_date")
        }),
        ("Scouting Work", {
            "fields": ("experience", "types", "competitions", "age_groups", "regions", "players_assessed", "shortlisted", "events")
        }),
        ("Football Background", {
            "fields": ("playing",)
        }),
        ("Scouting Skills and Technical Capabilities", {
            "fields": ("strengths",)
        }),
        ("Achievements and Recognition", {
            "fields": ("achievements",)
        }),
    )

    def thumb(self, obj):
        if obj.photo:
            return format_html('<img src="{}" style="height:42px;width:42px;object-fit:contain;border-radius:6px;background:#fff;" />', obj.photo.url)
        return "-"

    def preview(self, obj):
        if obj.photo:
            return format_html('<img src="{}" style="max-width:240px;max-height:260px;object-fit:contain;border-radius:10px;background:#fff;" />', obj.photo.url)
        return "No image"







@admin.register(MasterLabels)
class MasterLabelsAdmin(admin.ModelAdmin):
    list_display = ("id", "keydata", "lang")


def export_as_csv(self, request, queryset):

    meta = self.model._meta
    field_names = [field.name for field in meta.fields]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename={}.csv'.format(
        meta)
    writer = csv.writer(response)

    writer.writerow(field_names)
    for obj in queryset:
        row = writer.writerow([getattr(obj, field) for field in field_names])

    return response


admin.site.add_action(export_as_csv)

from django.contrib import admin
from .models import Career360Partner

@admin.register(Career360Partner)
class Career360PartnerAdmin(admin.ModelAdmin):
    list_display = ("name", "order", "is_active")
    list_editable = ("order", "is_active")
    search_fields = ("name",)

@admin.register(Lang)
class LangAdmin(admin.ModelAdmin):
    list_display = ('id', 'lang')
    search_fields = ('id', 'lang',)



    


from django.contrib import admin
from .models import FAQItem

@admin.register(FAQItem)
class FAQItemAdmin(admin.ModelAdmin):
    list_display = ("question", "page", "language", "order", "is_active", "updated_at")
    list_filter  = ("page", "language", "is_active")
    search_fields = ("question", "answer")
    ordering = ("order", "id")
    list_editable = ("order", "is_active")

    fields = (
        "page",
        "language",
        "question",
        "answer",
        "order",
        "is_active",
    )

from django.contrib import admin
from django.utils.html import format_html
from django.conf import settings
from .models import MessageWeblink

@admin.register(MessageWeblink)
class MessageWeblinkAdmin(admin.ModelAdmin):
    list_display = ("code", "heading", "view_page", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = (
        "code",
        "heading",
        "heading_hi",
        "message_en",
        "message_hi",
        "outro_en",
        "outro_hi",
    )
    readonly_fields = ("updated_at",)

    def view_page(self, obj):
        base = getattr(settings, "PUBLIC_SITE_URL", "").rstrip("/")
        url = obj.get_absolute_url()  # e.g. "/smseast/"
        full = f"{base}{url}" if base else url
        return format_html('<a href="{}" target="_blank">View page</a>', full)

    view_page.short_description = "Public link"

    # Optional: show "View on site" button on edit page too
    view_on_site = True

from django.contrib import admin
from django.utils.html import format_html
from django.conf import settings
from .models import ConsentNocWeblink


@admin.register(ConsentNocWeblink)
class ConsentNocWeblinkAdmin(admin.ModelAdmin):
    list_display = ("code", "title", "view_page", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("code", "title", "event_name", "venue_value", "date_value")
    readonly_fields = ("updated_at",)

    fieldsets = (
        ("URL + Status", {
            "fields": ("code", "is_active", "updated_at")
        }),
        ("Branding", {
            "fields": ("watermark_logo_url",)
        }),
        ("Top Content", {
            "fields": ("title", "event_name", "venue_label", "venue_value", "date_label", "date_value", "intro_text")
        }),
        ("Section Headings", {
            "fields": ("player_details_heading", "declaration_heading", "guardian_heading")
        }),
        ("Player Details Fields", {
            "fields": ("player_fields",)
        }),
        ("Declaration", {
            "fields": ("declaration_lead", "declaration_points")
        }),
        ("Guardian + Signature", {
            "fields": ("guardian_text", "signature_fields")
        }),
    )

    def view_page(self, obj):
        base = getattr(settings, "PUBLIC_SITE_URL", "").rstrip("/")
        url = obj.get_absolute_url()  # "/noc/<code>/"
        full = f"{base}{url}" if base else url
        return format_html('<a href="{}" target="_blank">View NOC</a>', full)

    view_page.short_description = "Public link"
    view_on_site = True





from django.contrib import admin
from .models import PageContent, PageImage, PartnerInterest


class PageImageInline(admin.TabularInline):
    model = PageImage
    extra = 1
    fields = ("image", "caption", "is_active")

from django.contrib import admin
from .models import StateAssociationPartner, StateAssociationPartnerStat

class StateAssociationPartnerStatInline(admin.TabularInline):
    model = StateAssociationPartnerStat
    extra = 1

@admin.register(StateAssociationPartner)
class StateAssociationPartnerAdmin(admin.ModelAdmin):
    list_display = ("name", "state_name", "lang", "is_active", "sort_order")
    list_filter = ("lang", "is_active")
    search_fields = ("name", "state_name")
    prepopulated_fields = {"slug": ("name",)}
    inlines = [StateAssociationPartnerStatInline]
from django.contrib import admin
from django.utils.html import format_html
from .models import ScoutOnWheel, ScoutOnWheelImage, ScoutOnWheelMediaArticle, ScoutOnWheelAndhraReportItem


class ScoutOnWheelImageInline(admin.TabularInline):
    model = ScoutOnWheelImage
    extra = 1
    fields = ("preview", "image", "caption", "sort_order", "is_active")
    readonly_fields = ("preview",)
    ordering = ("sort_order", "-created_at")
    verbose_name = "Glimpses from the initiative"
    verbose_name_plural = "Glimpses from the initiative"

    def preview(self, obj):
        if obj and obj.image:
            return format_html(
                '<img src="{}" style="height:48px;width:72px;object-fit:cover;border-radius:8px;" />',
                obj.image.url
            )
        return "—"


class ScoutOnWheelMediaInline(admin.TabularInline):
    model = ScoutOnWheelMediaArticle
    extra = 1
    fields = ("title", "source", "link", "cover_image", "sort_order", "is_active")
    ordering = ("sort_order", "-created_at")
    verbose_name = "IKF in News"
    verbose_name_plural = "IKF in News"


class ScoutOnWheelAndhraReportInline(admin.TabularInline):
    model = ScoutOnWheelAndhraReportItem
    extra = 1
    fields = ("title", "cover_image", "pdf_file", "link", "sort_order", "is_active")
    ordering = ("sort_order", "-created_at")
    verbose_name = "Scout on Wheel Andhra Pradesh Project"
# invictus/admin.py

from django.contrib import admin
from .models import TeamMember, Thought
from django.contrib import admin
from .models import TeamMember, Thought, InvictusCarouselImage

@admin.register(InvictusCarouselImage)
class InvictusCarouselImageAdmin(admin.ModelAdmin):
    list_display = ('__str__', 'is_active', 'order', 'created_at')
    list_filter = ('is_active',)
    search_fields = ('title', 'alt_text')
    ordering = ('order', 'id')
    list_editable = ('is_active', 'order')

@admin.register(TeamMember)
class TeamMemberAdmin(admin.ModelAdmin):
    list_display = ('name', 'designation', 'category', 'order')
    list_filter = ('category',)
    ordering = ('order',)


@admin.register(Thought)
class ThoughtAdmin(admin.ModelAdmin):
    list_display = ('name', 'designation', 'order')
    ordering = ('order',)

import csv
import io

from django.contrib import admin, messages
from django.urls import path
from django.shortcuts import render, redirect
from django.db import transaction

from django.contrib import admin
from .models import SeasonReport, PlayerEntry

@admin.register(SeasonReport)
class SeasonReportAdmin(admin.ModelAdmin):
    list_display = ("season", "title", "year_range", "cities", "participants", "shortlisted_count", "scouted_count", "international_scholarship_count", "clubs", "show_on_ikf_trials")
    list_editable = ("show_on_ikf_trials",)
    search_fields = ("title", "season", "year_range")
    fieldsets = (
        ("Season identity", {
            "fields": ("season", "title", "year_range", "description"),
        }),
        ("Stats (used on both IKF Trials and Scouted Players pages)", {
            "fields": (
                "cities",
                "participants",
                "shortlisted_count",
                "scouted_count",
                "international_scholarship_count",
                "clubs",
            ),
        }),
        ("Scouted Players page", {
            "fields": ("remarks",),
            "description": "Optional note shown above the player tables on the Scouted Players page.",
        }),
        ("IKF Trials visibility", {
            "fields": ("show_on_ikf_trials",),
        }),
    )



@admin.register(PlayerEntry)
class PlayerEntryAdmin(admin.ModelAdmin):
    list_display = ("season", "list_type", "sr_no", "player_name", "clubs")
    list_filter = ("season", "list_type")
    search_fields = ("player_name", "clubs")

    change_list_template = "admin/playerentry_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("upload-csv/", self.admin_site.admin_view(self.upload_csv), name="playerentry_upload_csv"),
        ]
        return custom_urls + urls

    def upload_csv(self, request):
        if request.method == "POST" and request.FILES.get("csv_file"):
            csv_file = request.FILES["csv_file"]

            if not csv_file.name.endswith(".csv"):
                self.message_user(request, "Please upload a .csv file", level=messages.ERROR)
                return redirect("..")

            data = csv_file.read().decode("utf-8", errors="ignore")
            f = io.StringIO(data)

# Peek first line to decide if it's a header or a title line
            first_line = f.readline().lstrip("\ufeff")  # removes BOM if any

# If first line looks like a header, rewind back to start
            if "Sr No" in first_line and "Players Name" in first_line:
               f.seek(0)

            reader = csv.DictReader(f)

# Normalize header keys (handles "Season " with trailing space)
            reader.fieldnames = [h.strip() for h in (reader.fieldnames or [])]


            # Normalize header keys by stripping spaces like "Season "
            reader.fieldnames = [h.strip() for h in (reader.fieldnames or [])]

            required = {"Sr No", "Players Name", "Clubs", "Season", "Shortlisted/selected"}
            if not required.issubset(set(reader.fieldnames or [])):
                self.message_user(
                    request,
                    f"Missing headers. Found: {reader.fieldnames}. Required: {sorted(required)}",
                    level=messages.ERROR,
                )
                return redirect("..")

            created, updated, skipped = 0, 0, 0

            with transaction.atomic():
                for row in reader:
                    # Clean keys (DictReader already uses normalized fieldnames)
                    sr_no_raw = (row.get("Sr No") or "").strip()
                    name = (row.get("Players Name") or "").strip()
                    clubs = (row.get("Clubs") or "").strip()
                    season_raw = (row.get("Season") or "").strip()
                    lt_raw = (row.get("Shortlisted/selected") or "").strip()

                    # skip empty rows
                    if not (sr_no_raw and name and season_raw and lt_raw):
                        skipped += 1
                        continue

                    try:
                        sr_no = int(sr_no_raw)
                        season_num = int(season_raw)
                    except ValueError:
                        skipped += 1
                        continue

                    # Normalize list type
                    lt_lower = lt_raw.lower()
                    if lt_lower in ["shortlisted", "shortlisted players"]:
                        list_type = "Shortlisted"
                    elif lt_lower in ["selected", "scouted", "scouted players"]:
                        list_type = "Scouted"
                    else:
                        # fallback: treat unknown as skipped
                        skipped += 1
                        continue

                    # Ensure season exists (so season_id is never null)
                    season_obj, _ = SeasonReport.objects.get_or_create(
                        season=season_num,
                        defaults={"title": f"Tyger IKF Season {season_num} Players Report"},
                    )

                    obj, was_created = PlayerEntry.objects.update_or_create(
                        season=season_obj,
                        list_type=list_type,
                        sr_no=sr_no,
                        defaults={"player_name": name, "clubs": clubs},
                    )

                    if was_created:
                        created += 1
                    else:
                        updated += 1

            self.message_user(
                request,
                f"CSV processed ✅ Created: {created}, Updated: {updated}, Skipped: {skipped}",
                level=messages.SUCCESS,
            )
            return redirect("..")

        # GET → show upload form
        context = dict(self.admin_site.each_context(request))
        return render(request, "admin/upload_csv.html", context)

@admin.register(ScoutOnWheel)
class ScoutOnWheelAdmin(admin.ModelAdmin):
    list_display = ("title", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("title",)
    inlines = [ScoutOnWheelImageInline, ScoutOnWheelMediaInline, ScoutOnWheelAndhraReportInline]


from .models import NaariShakti, NaariShaktiImage, NaariShaktiImpactStat


class NaariShaktiImageInline(admin.TabularInline):
    model = NaariShaktiImage
    extra = 1
    fields = ("preview", "image", "caption", "sort_order", "is_active")
    readonly_fields = ("preview",)
    ordering = ("sort_order", "-created_at")
    verbose_name = "Glimpses from the initiative"
    verbose_name_plural = "Glimpses from the initiative"

    def preview(self, obj):
        if obj and obj.image:
            return format_html(
                '<img src="{}" style="height:48px;width:72px;object-fit:cover;border-radius:8px;" />',
                obj.image.url
            )
        return "—"


class NaariShaktiImpactStatInline(admin.TabularInline):
    model = NaariShaktiImpactStat
    extra = 1
    fields = ("value", "title", "description", "sort_order", "is_active")
    ordering = ("sort_order", "id")
    verbose_name = "Impact So Far — stat"
    verbose_name_plural = "Impact So Far"


@admin.register(NaariShakti)
class NaariShaktiAdmin(admin.ModelAdmin):
    list_display = ("title", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("title",)
    inlines = [NaariShaktiImpactStatInline, NaariShaktiImageInline]


# ─────────────────────────────────────────────────────────────
# IKF 360 Pathway Partners — admin for logos on /ikf360vision
# ─────────────────────────────────────────────────────────────
from .models import IKF360PathwayPartner, ScoutOnWheels2026StateMap


@admin.register(IKF360PathwayPartner)
class IKF360PathwayPartnerAdmin(admin.ModelAdmin):
    list_display = ("preview", "name", "category", "website", "sort_order", "is_active", "updated_at")
    list_display_links = ("preview", "name")
    list_filter = ("is_active",)
    search_fields = ("name", "category", "website")
    list_editable = ("sort_order", "is_active")
    ordering = ("sort_order", "-created_at")
    readonly_fields = ("preview_large", "created_at", "updated_at")

    fieldsets = (
        ("Partner", {
            "fields": ("name", "category", "website"),
            "description": "‘Category’ is an optional label shown below the logo on the "
                           "page, e.g. “Nudge Sports” or “Health Partner”.",
        }),
        ("Logo", {
            "description": (
                "<b>Recommended logo size:</b> 400 × 200 px (2 : 1 ratio).<br>"
                "PNG with a transparent background looks best. "
                "Min width 300 px, max width 800 px. "
                "Keep file size under 200 KB."
            ),
            "fields": ("logo", "preview_large"),
        }),
        ("Display", {
            "fields": ("sort_order", "is_active"),
        }),
        ("Meta", {
            "classes": ("collapse",),
            "fields": ("created_at", "updated_at"),
        }),
    )

    def preview(self, obj):
        if obj and obj.logo:
            return format_html(
                '<img src="{}" style="height:36px;max-width:90px;object-fit:contain;'
                'background:#fafafa;border:1px solid #eee;border-radius:6px;padding:3px;" />',
                obj.logo.url,
            )
        return "—"
    preview.short_description = "Logo"

    def preview_large(self, obj):
        if obj and obj.logo:
            return format_html(
                '<img src="{}" style="max-height:160px;max-width:340px;object-fit:contain;'
                'background:#fafafa;border:1px solid #eee;border-radius:8px;padding:8px;" />',
                obj.logo.url,
            )
        return "Upload a logo to see a preview."
    preview_large.short_description = "Current logo preview"


# ─────────────────────────────────────────────────────────────
# Scout on Wheels 2026 — admin for state route maps
# ─────────────────────────────────────────────────────────────
@admin.register(ScoutOnWheels2026StateMap)
class ScoutOnWheels2026StateMapAdmin(admin.ModelAdmin):
    list_display = ("preview", "display_title_admin", "state", "sort_order", "is_active", "updated_at")
    list_display_links = ("preview", "display_title_admin")
    list_filter = ("state", "is_active")
    search_fields = ("title", "state")
    list_editable = ("sort_order", "is_active")
    ordering = ("sort_order", "-created_at")
    readonly_fields = ("preview_large", "created_at", "updated_at")

    fieldsets = (
        ("State", {
            "fields": ("state", "title"),
            "description": (
                "Pick one of the 3 states. <b>Title</b> is optional — "
                "leave blank to show the state name on the card."
            ),
        }),
        ("Route Map Image", {
            "description": (
                "<b>Recommended size:</b> 1200 × 800 px (3 : 2 ratio).<br>"
                "PNG or JPG. The card thumbnail is cropped to a fixed shape; "
                "<b>clicking the card opens the full image at its original aspect ratio</b>, "
                "so upload the route map at its true ratio. "
                "Min width 800 px, max width 2000 px. Keep file size under 500 KB."
            ),
            "fields": ("route_map_image", "preview_large"),
        }),
        ("Display", {
            "fields": ("sort_order", "is_active"),
        }),
        ("Meta", {
            "classes": ("collapse",),
            "fields": ("created_at", "updated_at"),
        }),
    )

    def display_title_admin(self, obj):
        return obj.display_title
    display_title_admin.short_description = "Card title"

    def preview(self, obj):
        if obj and obj.route_map_image:
            return format_html(
                '<img src="{}" style="height:48px;width:72px;object-fit:cover;'
                'border-radius:6px;border:1px solid #eee;" />',
                obj.route_map_image.url,
            )
        return "—"
    preview.short_description = "Map"

    def preview_large(self, obj):
        if obj and obj.route_map_image:
            return format_html(
                '<img src="{}" style="max-height:260px;max-width:420px;object-fit:contain;'
                'background:#fafafa;border:1px solid #eee;border-radius:8px;padding:6px;" />',
                obj.route_map_image.url,
            )
        return "Upload a route map to see a preview."
    preview_large.short_description = "Current map preview"


# ─────────────────────────────────────────────────────────────
# Scouted Players Roasters — admin for the 2 PDF cover cards
# ─────────────────────────────────────────────────────────────
from .models import ScoutedPlayersRoaster


@admin.register(ScoutedPlayersRoaster)
class ScoutedPlayersRoasterAdmin(admin.ModelAdmin):
    list_display = (
        "preview",
        "title",
        "roaster_type",
        "has_pdf",
        "sort_order",
        "is_active",
        "updated_at",
    )
    list_display_links = ("preview", "title")
    list_filter = ("roaster_type", "is_active")
    search_fields = ("title", "subtitle")
    list_editable = ("sort_order", "is_active")
    ordering = ("sort_order", "-created_at")
    readonly_fields = ("preview_large", "created_at", "updated_at")

    fieldsets = (
        ("Roaster", {
            "fields": ("roaster_type", "title", "subtitle"),
            "description": (
                "Pick whether this card is the National or International roaster. "
                "Only one card per type is allowed."
            ),
        }),
        ("Cover image", {
            "description": (
                "<b>Recommended size:</b> 800 × 1131 px (A4 portrait, ratio 1 : √2).<br>"
                "PNG or JPG. The card crops the cover to a uniform shape — "
                "clicking the card opens the full PDF in a new tab. "
                "Keep file size under 400 KB."
            ),
            "fields": ("cover_image", "preview_large"),
        }),
        ("PDF source", {
            "description": (
                "Use <b>PDF file</b> if your PDF is under ~50 MB. "
                "For very large files (the National roaster is ~336 MB and the "
                "International roaster is ~65 MB), host the PDF on Google Drive / "
                "S3 / Dropbox and paste the public link into <b>PDF external URL</b>. "
                "If both are set, the uploaded file wins."
            ),
            "fields": ("pdf_file", "pdf_external_url"),
        }),
        ("Display", {
            "fields": ("sort_order", "is_active"),
        }),
        ("Meta", {
            "classes": ("collapse",),
            "fields": ("created_at", "updated_at"),
        }),
    )

    def preview(self, obj):
        if obj and obj.cover_image:
            return format_html(
                '<img src="{}" style="height:64px;width:48px;object-fit:cover;'
                'border-radius:4px;border:1px solid #eee;" />',
                obj.cover_image.url,
            )
        return "—"
    preview.short_description = "Cover"

    def preview_large(self, obj):
        if obj and obj.cover_image:
            return format_html(
                '<img src="{}" style="max-height:320px;max-width:240px;object-fit:contain;'
                'background:#fafafa;border:1px solid #eee;border-radius:8px;padding:6px;" />',
                obj.cover_image.url,
            )
        return "Upload a cover image to see a preview."
    preview_large.short_description = "Current cover preview"

    def has_pdf(self, obj):
        if obj.pdf_file:
            return format_html('<span style="color:#16a34a;font-weight:700;">{}</span>', '✓ uploaded')
        if obj.pdf_external_url:
            return format_html('<span style="color:#2563eb;font-weight:700;">{}</span>', '✓ external link')
        return format_html('<span style="color:#ef4444;font-weight:700;">{}</span>', '✗ none')
    has_pdf.short_description = "PDF"

@admin.register(PageContent)
class PageContentAdmin(admin.ModelAdmin):
    # 🔹 removed admin_label from list_display
    # and show a human-readable name from the choices instead
    list_display = ("page_key", "page_name")
    list_filter = ("page_key",)
    search_fields = ("page_key",)
    inlines = [PageImageInline]

    def page_name(self, obj):
        # uses the label from PAGE_CHOICES (e.g. "IKF Trials")
        return obj.get_page_key_display()
    page_name.short_description = "Page"
    
    exclude = ("admin_label",)

@admin.register(PartnerInterest)
class PartnerInterestAdmin(admin.ModelAdmin):
    list_display  = ("org_name","contact_person","email","city","focus_areas","created_at")
    search_fields = ("org_name","contact_person","email","city","focus_areas")
    list_filter   = ("created_at",)



from .models import ScoutOnWheelsMedia

class ScoutOnWheelsMediaInline(admin.TabularInline):
    model = ScoutOnWheelsMedia
    extra = 0
    fields = ("title", "youtube_url_or_id", "is_active", "sort_order", "created_at")
    readonly_fields = ("created_at",)
    ordering = ("sort_order",)



from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse
from django.conf import settings
from .models import RepTemplate

@admin.action(description="Export onboarding links (CSV)")
def export_onboarding_links(modeladmin, request, queryset):
    import csv
    from django.http import HttpResponse
    base = getattr(settings, "SITE_BASE_URL", request.build_absolute_uri("/").rstrip("/"))
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = 'attachment; filename="rep_onboarding_links.csv"'
    w = csv.writer(resp)
    w.writerow(["rep_name", "rep_num", "onboarding_token", "onboarding_url"])
    for rep in queryset:
        path = reverse("rep_onboarding", args=[rep.onboarding_token])
        w.writerow([rep.rep_name, rep.rep_num, rep.onboarding_token, f"{base}{path}"])
    return resp




@admin.register(HomeImages)
class HomeImagesAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Navbar)
class NavbarAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'urlitem', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(HomeBanner)
class HomeBannerAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'heading_1', 'heading_2_colored', 'description', 'button1_text', 'button2_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(AboutUs)
class AboutUsAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic_1', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Winners)
class WinnersAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'name', 'pic', 'category', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Partners)
class PartnersAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'pic', 'description', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Initiatives)
class InitiativesAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'name', 'pic', 'description', 'initiative_href', 'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(MoreaboutIKF)
class moreaboutIKFAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'name', 'pic', 'description', 'initiative_href', 'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Buttons_IKF_FOR_ALL)
class Buttons_IKF_FOR_ALLAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Photos_IKF_FOR_ALL)
class Photos_IKF_FOR_ALLAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'relationkey', 'relationkeysecondary', 'relationkeytertiary', 'relationkeyquaternary', 'relationkeyfifth', 'relationkeysixth', 'size', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(News)
class NewsAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'heading', 'description', 'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4', 'publisher', 'priority', 'type'
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(WorkshopNews)
class WorkshopNewsAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'heading', 'description', 'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Tab)
class TabAdmin(admin.ModelAdmin):
    list_display = ('id', 'tab', 'extra', )


@admin.register(SubTab)
class SubTabAdmin(admin.ModelAdmin):
    list_display = ('id', 'subtab', 'tab', )


@admin.register(AboutTrials_BannerPhoto)
class AboutTrials_BannerPhoto(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'lang')
    search_fields = ('id', 'keydata', 'name')


@admin.register(ActiveStatusNav)
class ActiveStatusNavAdmin(admin.ModelAdmin):
    list_display = ('id',)
    search_fields = ('id', )


@admin.register(TrialsAndInitiativeType)
class TrialsAndInitiativeTypeAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata')
    search_fields = ('id', 'keydata',)


@admin.register(TrialsAndInitiativeNav)
class TrialsAndInitiativeNavAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'pic', 'name', 'href', 'description', 'bottomtext', 'listtext', 'pagetype', 'activestatus', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    list_filter = ('pagetype',)
    search_fields = ('id', 'keydata', 'name')


@admin.register(AboutTrials_Feature)
class AboutTrials_Feature(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'description', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Season1_Feature)
class Season1_Feature(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'description', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Season1_Tabs)
class Season1_Tabs(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'name', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Season1_Bottom)
class Season1_Buttom(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'description', 'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(About_section1)
class About_section1(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading1', 'paragraph1', 'heading2', 'paragraph2', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(TeamVideoCarousel)
class TeamVideoCarousel(admin.ModelAdmin):
    list_display = ('id',  'about' , 'video_url'
                    )
    search_fields = ('id', 'name')
@admin.register(About_TeamAdvisors)
class About_Team1(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'post', 'about', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(About_TeamCore_Team)
class About_TeamCore_Team(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'post', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',)
    search_fields = ('id', 'keydata', 'name')


@admin.register(About_OURTEAM_of_IkF)
class About_TeamSupporters_of_IkF(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'post', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(IKF_Tech1)
class IKF_Tech1(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'heading', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(IKF_Tech2)
class IKF_Tech2(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'heading', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

from django.contrib import admin
from .models import ContactUs

@admin.register(ContactUs)
class ContactUsAdmin(admin.ModelAdmin):
    list_display = ('name', 'email', 'mobile', 'message', 'created_at')
    search_fields = ('name', 'email', 'mobile', 'subject', 'message')
    list_filter = ('created_at',)
    ordering = ('-created_at',)

@admin.register(IKF_Tech3)
class IKF_Tech3(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(IKF_Tech4)
class IKF_Tech4(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'heading', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Regional_Partners1)
class Regional_Partners1(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'heading', 'paragraph', 'button1_text', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(National_Partners1)
class National_Partners1(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'heading', 'paragraph', 'button1_text', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(International_Partners1)
class International_Partners1(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'heading', 'paragraph', 'button1_text', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Corporate_Partners1)
class Corporate_Partners1(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'heading', 'paragraph', 'button1_text', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Clubs_Partners1)
class Clubs_Partners1(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'heading', 'paragraph', 'button1_text', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(WorkshopInternationalLogo)
class WorkshopInternationalLogoAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'heading', 'paragraph', 'button1_text', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')



@admin.register(Playabroad)
class playabroad(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'button_text', 'button_href', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TabelWorkshops)
class TabelWorkshops(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'First_Column', 'Second_Column', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(TabelPastTrial)
class TabelPastTrial(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'First_Column', 'Second_Column', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')
@admin.register(WorkShopsButton)
class WorkShopsButton(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(WorkShopsFeatureStrip)
class WorkShopsFeatureStrip(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading', 'paragraph', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(WorkShopsExperts)
class WorkShopsExperts(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading1', 'heading2', 'button_text', 'paragraph', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(AboutTrials_Images)
class AboutTrials_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Season1_Images)
class Season1_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Season2_Images)
class Season2_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(ScoutingCertification_Images)
class ScoutingCertification_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(PlayAbroad_Images)
class PlayAbroad_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(WorkShops_Images)
class WorkShops_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(IKF_Technology_Images)
class IKF_Technology_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(About_Us_Images)
class About_Us_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Reginoal_Images)
class Reginoal_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(National_Images)
class National_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(International_Images)
class International_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Corporate_Images)
class Corporate_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Clubs)
class Clubs(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Season1_Button)
class Season1_Button(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'button_text', 'button_href', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Associate_Corporate_Images)
class Associate_Corporate_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(HomePage)
class HomePage(admin.ModelAdmin):
    list_display = ('id',  'image', 'en', 'marathi', 'hindi',
                    )
    search_fields = ('id', 'name')
@admin.register(Associate_Philanthropists_Images)
class Associate_Philanthropists_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Associate_PSUS_Images)
class Associate_PSUS_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Associate_Academy_Images)
class Associate_Academy_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Associate_Corporate_Main)
class Associate_Corporate_Main(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading1', 'paragraph1', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


from django.contrib import admin
from .models import TrialSchedule, TrialDay

class TrialDayInline(admin.TabularInline):  # or use StackedInline for full-width fields
    model = TrialDay
    extra = 1  # how many empty forms to show
    fields = ('order', 'day_title', 'description')
    ordering = ('order',)

@admin.register(TrialSchedule)
class TrialScheduleAdmin(admin.ModelAdmin):
    list_display = ('city', 'updated_at')
    inlines = [TrialDayInline]
    search_fields = ('city',)
    ordering = ('city',)


from django.contrib import admin
from .models import HomePageClub

@admin.register(HomePageClub)
class HomePageClubAdmin(admin.ModelAdmin):
    list_display = ('name', 'pic')

from django.contrib import admin
from .models import WorkshopLocation

@admin.register(WorkshopLocation)
class WorkshopLocationAdmin(admin.ModelAdmin):
    list_display = ('city', 'dates', 'venue', 'maps_link', 'price', 'priority')
    search_fields = ('city', 'venue')
    ordering = ('priority',)


@admin.register(Associate_Philanthropists_Main)
class Associate_Philanthropists_Main(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading1', 'paragraph1', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Associate_PSUS_Main)
class Associate_PSUS_Main(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading1', 'paragraph1', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Associate_Corporate_Right)
class Associate_Corporate_Right(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading2', 'paragraph2', 'button_text', 'href', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Associate_Philanthropists_Right)
class Associate_Philanthropists_Right(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading2', 'paragraph2', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Associate_PSUS_Right)
class Associate_PSUS_Right(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading2', 'paragraph2', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(UpCommingMatchs)
class UpCommingMatchs(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'teamname_1st', 'teamname_2nd', 'include', 'Match_Month', 'day_remaining', 'date_year', 'Match_Time', 'Match_City', 'Team1_pic', 'Team2_pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4', 'matchdate'
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(MasterState)
class MasterStateAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'include',)
    search_fields = ('id', 'name',)


@admin.register(MasterCity)
class MasterCityAdmin(admin.ModelAdmin):
    list_display = ('id', 'city', 'state', 'include')
    search_fields = ('id', 'city',)


@admin.register(MasterPosition)
class MasterPositionAdmin(admin.ModelAdmin):
    list_display = ('id', 'position', 'label')
    search_fields = ('position', 'label',)


@admin.register(MasterIkfSeasonUi)
class MasterIkfSeasonUiAdmin(admin.ModelAdmin):
    list_display = ('id', 'name', 'include')
    search_fields = ('id', 'name')


@admin.register(MasterSeason)
class MasterSeasonAdmin(admin.ModelAdmin):
    list_display = ('id', 'en', 'year')
    search_fields = ('id',)


@admin.register(TabelSeason2)
class TabelSeason2(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'First_Column', 'Second_Column', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Season2_Features)
class Season2_Features(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'description', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TabelScoutingCertification)
class TabelScoutingCertification(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'First_Column', 'Second_Column', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Scouting_Features)
class Scouting_Features(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'description', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Scouting_Initiatives)
class Scouting_Initiatives(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'name', 'pic', 'description', 'initiative_href', 'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TabelPlayabroad)
class TabelPlayabroad(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'First_Column', 'Second_Column', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Playabroad_Features)
class Playabroad_Features(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'description', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Playabroad_Initiatives)
class Playabroad_Initiatives(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'name', 'pic', 'description', 'initiative_href', 'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(NavTabs)
class Playabroad_Initiatives(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'tab', 'subtab', 'tab_href', 'tab_href', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay1_Corporate)
class TakeaWay1_Corporate(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay2_Corporate)
class TakeaWay2_Corporate(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay2_Philanthropists)
class TakeaWay2_Philanthropists(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay1_Philanthropists)
class TakeaWay1_Philanthropists(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay1_PSUS)
class TakeaWay1_PSUS(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay2_PSUS)
class TakeaWay2_PSUS(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay3_PSUS)
class TakeaWay3_PSUS(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TabelWorkshopsReg)
class TabelWorkshopsReg(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'First_Column', 'Second_Column', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(WorkShopsRegButton)
class WorkShopsRegButton(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(WorkShopsRegFeatureStrip)
class WorkShopsRegFeatureStrip(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading', 'paragraph', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(WorkShopsRegExperts)
class WorkShopsRegExperts(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading1', 'heading2', 'button_text', 'paragraph', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(WorkShopsReg_Images)
class WorkShopsReg_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Selected_Players_Season1)
class Selected_Players_Season1(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'Name', 'Age', 'Information', 'pic', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Selected_Players_Season2)
class Selected_Players_Season2(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'Name', 'Age', 'Information', 'pic', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay1_Academy)
class TakeaWay1_Academy(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay2_Academy)
class TakeaWay2_Academy(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Associate_Clubs_Images)
class Associate_Clubs_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Associate_TournamentOrganizer_Images)
class Associate_TournamentOrganizer_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Associate_Clubs_Right)
class Associate_Clubs_Right(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading2', 'paragraph2', 'button_text', 'href', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(TakeaWay1_Clubs)
class TakeaWay1_Clubs(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay2_Clubs)
class TakeaWay2_Clubs(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(TakeaWay3_Clubs)
class TakeaWay3_Clubs(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Clubs_Parter_Home)
class Clubs_Partner_HomeAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4', 'is_international','is_participated')
    search_fields = ('id', 'keydata', 'name')


@admin.register(Clubs_Parter_Season1)
class Clubs_Partner_Season1Admin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Clubs_Parter_Season2)
class Clubs_Partner_Season2Admin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Season3_Images)
class Season3_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TabelSeason3)
class TabelSeason3(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'First_Column', 'Second_Column', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(Season3_Features)
class Season3_Features(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'description', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay1_TournamentOrganizer)
class TakeaWay1_TournamentOrganizer(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')


@admin.register(TakeaWay2_TournamentOrganizer)
class TakeaWay2_TournamentOrganizer(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'paragraph', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Associate_TournamentOrganizer_Right)
class Associate_TournamentOrganizer_Right(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading2', 'paragraph2', 'button_text', 'href', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Previous_Workshop_Images)
class Previous_Workshop_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Winners_Workshop_Images)
class Winners_Workshop_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Testimonials)
class Testimonials(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic','description', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Mentor)
class Mentor(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic','description',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Clubs_Parter_abouttrials)
class Clubs_Parter_abouttrials(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Youtube)
class YoutubeAdmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'heading', 'description', 'button_text', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(UserFlow_Season3)
class UserFlow_Season3(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading1', 'paragraph1', 'heading2', 'paragraph2', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Calendar_Season3)
class Calendar_Season3(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'Description', 'Month','lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Winners_Season2)
class Winners_Season2Admin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'name', 'pic', 'category', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(InstaFeed)
class InstaFeed(admin.ModelAdmin):
    list_display = ('id', 'keydata','Link','lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Photo_Gallery_Images)
class Photo_Gallery_Images(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'size', 'pic', 'name', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Gallery_Buttons)
class Gallery_Buttons(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'BCity', 'Bname','BBehavior','Image', 'lang',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(Resources)
class Resourcesadmin(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'document', 'lang', 'link', 'description', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(SupportedByFinancial)
class SupportedByFinancialAdmin(admin.ModelAdmin):
    list_display = ('id', 'financial_year', 'order_by')
    search_fields = ('id','financial_year',)

@admin.register(SupportedBy)
class SupportedByAdmin(admin.ModelAdmin):
    list_display = ('id', 'logo_pic', 'company_name', 'company_address', 'project_brief', 'partners', 'button_link', 'order_by', 'button_text', 'attr1', 'attr2', 'attr3', 'attr4', 'attr5')
    search_fields = ('id', 'company_name', 'company_address',)

@admin.register(SupportedByImages)
class SupportedByImagesAdmin(admin.ModelAdmin):
    list_display = ('id', 'pic',)

@admin.register(SupportedByImages2526)
class SupportedByImages2526Admin(admin.ModelAdmin):
    list_display = ('id', 'pic', 'order_by')
    ordering = ('order_by',)
from django.utils.html import format_html

@admin.register(Donate_Images)
class DonateImagesAdmin(admin.ModelAdmin):
    list_display = ('image_preview',)  # ✅ Only show image preview in the list
    readonly_fields = ('image_preview',)

    def image_preview(self, obj):
        if obj.pic:
            return format_html('<img loading="lazy" src="{}" style="height:100px;" />', obj.pic.url)
        return "No image uploaded"

    image_preview.short_description = 'Image'

from django.contrib import admin
from django.urls import reverse
from django.utils.html import format_html

from .models import (
    NoticeboardBox, NoticeboardNotice, NoticeBoardInfoBlock,
    TrialGuidelinesPage, TrialGuidelinesSection, TrialGuidelinesItem
)


@admin.register(NoticeBoardInfoBlock)
class NoticeBoardInfoBlockAdmin(admin.ModelAdmin):
    list_display = ("title", "order", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("title", "subtitle", "body")
    ordering = ("order", "-created_at")

    fieldsets = (
        ("Content", {"fields": ("title", "subtitle", "body", "image")}),
        ("CTA (optional)", {"fields": ("cta_text", "cta_url")}),
        ("Visibility & Ordering", {"fields": ("is_active", "order")}),
    )


from django.utils.text import slugify

@admin.register(NoticeboardBox)
class NoticeboardBoxAdmin(admin.ModelAdmin):
    list_display = ("name", "slug", "url", "order", "is_active", "view_page")
    list_editable = ("order", "is_active")
    search_fields = ("name", "slug", "url")
    prepopulated_fields = {"slug": ("name",)}

    def save_model(self, request, obj, form, change):
        # Auto-generate slug with no hyphens or spaces
        if obj.name:
            obj.slug = slugify(obj.name).replace("-", "")
        super().save_model(request, obj, form, change)

    def view_page(self, obj):
        if obj.slug:
            public_url = reverse("noticeboard_slug", kwargs={"slug": obj.slug})
            return format_html('<a class="button" href="{}" target="_blank">View Page</a>', public_url)
        return "—"
    view_page.short_description = "View Page"


@admin.register(NoticeboardNotice)
class NoticeboardNoticeAdmin(admin.ModelAdmin):
    list_display = ("title", "published_date", "is_active", "created_at")
    list_filter = ("is_active", "published_date")
    search_fields = ("title", "meta", "url")
    ordering = ("-published_date", "-created_at")


# ─────────────────────────────────────────────────────────────
# Trial Guidelines Admin (simple for non-tech)
# - 1 admin heading for the page
# - sections editable in Page screen
# - items editable inside each Section screen
# ─────────────────────────────────────────────────────────────

class TrialGuidelinesSectionInline(admin.TabularInline):
    model = TrialGuidelinesSection
    extra = 1
    fields = ("order", "area", "title", "is_active")  # ✅ add area here
    ordering = ("area", "order",)  # ✅ keeps main + sidebar grouped nicely



@admin.register(TrialGuidelinesPage)
class TrialGuidelinesPageAdmin(admin.ModelAdmin):
    list_display = ("page_title", "slug", "is_active", "updated_at", "view_public_page")
    list_filter = ("is_active",)
    search_fields = ("page_title", "page_kicker", "slug")
    prepopulated_fields = {"slug": ("page_title",)}
    inlines = [TrialGuidelinesSectionInline]

    def view_public_page(self, obj):
        url = reverse("noticeboard_slug", kwargs={"slug": obj.slug})
        return format_html('<a href="{}" target="_blank">Open</a>', url)
    view_public_page.short_description = "Public Page"


class TrialGuidelinesItemInline(admin.TabularInline):
    model = TrialGuidelinesItem
    extra = 1
    fields = ("order", "item_type", "bold_label", "text", "is_active")
    ordering = ("order",)


@admin.register(TrialGuidelinesSection)
class TrialGuidelinesSectionAdmin(admin.ModelAdmin):
    list_display = ("title", "page", "order", "is_active")
    list_filter = ("is_active", "page")
    search_fields = ("title",)
    ordering = ("page", "order")
    inlines = [TrialGuidelinesItemInline]

# admin.py
from django.contrib import admin
from .models import PostTrialMediaPage, PostTrialMediaVideo

class PostTrialMediaVideoInline(admin.TabularInline):
    model = PostTrialMediaVideo
    extra = 1
    fields = ("order", "city_name", "video_url", "is_active")
    ordering = ("order",)

@admin.register(PostTrialMediaPage)
class PostTrialMediaPageAdmin(admin.ModelAdmin):
    list_display = ("title", "slug", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("title", "slug")
    inlines = [PostTrialMediaVideoInline]




@admin.register(TabelDonate)
class TabelDonateAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'keydata', 'First_Column', 'Second_Column', 'lang', 
        'attr1', 'attr2', 'attr3', 'attr4', 'image_tag'  # Show image preview
    )
    search_fields = ('id', 'keydata', 'First_Column', 'Second_Column')
    readonly_fields = ('image_tag',)

    def image_tag(self, obj):
        if obj.image:
            return format_html('<img loading="lazy" src="{}" width="60" height="60" style="object-fit: cover;" />', obj.image.url)
        return "-"
    image_tag.short_description = 'Image'




@admin.register(DonateFeatureStrip)
class DonateFeatureStrip(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading', 'paragraph', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')

@admin.register(PaymentFeatureStrip)
class PaymentFeatureStrip(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading', 'paragraph', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',
                    )
    search_fields = ('id', 'keydata', 'name')
from django.utils.html import format_html

from django.contrib import admin
from .models import SitemapCategory, SitemapItem


class SitemapItemInline(admin.TabularInline):
    model           = SitemapItem
    extra           = 1
    fields          = ("order", "title", "description", "url", "is_active")
    ordering        = ("order",)
    show_change_link = True


@admin.register(SitemapCategory)
class SitemapCategoryAdmin(admin.ModelAdmin):
    list_display    = ("name", "column", "order", "is_active", "item_count")
    list_editable   = ("column", "order", "is_active")
    search_fields   = ("name",)
    inlines         = [SitemapItemInline]

    @admin.display(description="Items")
    def item_count(self, obj):
        return obj.items.count()


@admin.register(SitemapItem)
class SitemapItemAdmin(admin.ModelAdmin):
    list_display    = ("title", "category", "order", "is_active", "short_url")
    list_editable   = ("order", "is_active")
    list_filter     = ("category", "is_active")
    search_fields   = ("title", "description", "url")
    ordering        = ("category__order", "order")

    fieldsets = (
        (None, {
            "fields": ("category", "title", "is_active", "order"),
        }),
        ("Content (shown in hover / tap preview)", {
            "fields": ("description", "url"),
            "description": (
                "The <strong>description</strong> appears in the yellow preview card "
                "when a visitor hovers over (desktop) or taps (mobile) the item.  "
                "The <strong>URL</strong> is where the &ldquo;Know More&rdquo; button leads."
            ),
        }),
    )

    @admin.display(description="URL")
    def short_url(self, obj):
        return obj.url[:60] + ("…" if len(obj.url) > 60 else "")
class NaariShaktiEditionFilter(admin.SimpleListFilter):
    title = 'Naari Shakti edition'
    parameter_name = 'naari_shakti_edition'

    def lookups(self, request, model_admin):
        return (
            ('bihar', 'IKF Naari Shakti - Bihar'),
            ('rajasthan', 'IKF Naari Shakti - Rajasthan'),
            ('meerut', 'IKF Naari Shakti - Meerut'),
            ('india', 'IKF Naari Shakti - India'),
        )

    def queryset(self, request, queryset):
        from django.db.models import Q
        val = self.value()
        if val == 'bihar':
            return queryset.filter(address__icontains='IKF Naari Shakti - Bihar Edition')
        if val == 'rajasthan':
            return queryset.filter(address__icontains='IKF Naari Shakti - Rajasthan')
        if val == 'meerut':
            return queryset.filter(address__icontains='IKF Naari Shakti - Meerut Edition')
        if val == 'india':
            return queryset.filter(
                Q(address__icontains='Naari Shakti') &
                ~Q(address__icontains='Bihar') &
                ~Q(address__icontains='Rajasthan') &
                ~Q(address__icontains='Meerut')
            )
        return queryset


@admin.register(Donation)
class DonationAdmin(admin.ModelAdmin):
    list_display = (
        'name',
        'email',
        'mobile',
        'amount',
        'status',
        'donated_for',          # single column showing the chosen project(s)
        'razorpay_order_id',
        'created_at',
    )
    list_select_related = ('project',)
    raw_id_fields = ('project',)

    def project_display(self, obj):
        """Show the admin-managed DonateProject name.
        Falls back to the project string parsed from the legacy `address`
        field for historical rows that don't have the FK set yet."""
        if obj.project_id and obj.project:
            return obj.project.name
        addr = (obj.address or '').replace('Support for:', '').strip()
        # First comma-separated value (most donations were single-project)
        first = addr.split(',')[0].strip() if addr else ''
        return first or '—'
    project_display.short_description = 'Project'
    project_display.admin_order_field = 'project__name'

    def donated_for(self, obj):
        """Show the project this donation was made for, always using the
        LIVE name from the DonateProject admin so renames are reflected.

        Priority:
          1. FK to DonateProject — live name (modern rows)
          2. Boolean flag → mapped slug → live DonateProject.name (legacy rows)
          3. Hardcoded label for the boolean (if slug isn't mapped/found)
          4. Legacy `address` string parsing (oldest rows)
        """
        # 1. Modern rows: direct FK
        if obj.project_id and obj.project:
            return obj.project.name

        # 2. Legacy boolean → live project name via slug cache
        slug_to_name = getattr(self, '_slug_to_name', {})
        names = []
        for field in self.purpose_fields:
            if not getattr(obj, field, False):
                continue
            slug = self.purpose_to_slug.get(field)
            if slug and slug in slug_to_name:
                names.append(slug_to_name[slug])
            else:
                # 3. Fall back to the hardcoded label for fields with no slug
                names.append(self.donation_purpose_labels.get(field, field))

        if names:
            # de-duplicate while preserving order (e.g. all 3 Field-of-Dreams
            # booleans collapse to one "IKF 360 - Field of Dreams" entry)
            seen, unique = set(), []
            for n in names:
                if n not in seen:
                    seen.add(n)
                    unique.append(n)
            return ', '.join(unique)

        # 4. Oldest fallback
        return self.project_display(obj)
    donated_for.short_description = 'Donated For'

    # Fallback labels — used only when the slug lookup misses (i.e. the
    # DonateProject row was deleted). Match the current project names in
    # the DonateProject admin so legacy data stays consistent.
    donation_purpose_labels = {
        'invictus_icare':   'I-Care by Invictus',
        'future_champions': 'IKF 360 - Field of Dreams',
        'books_and_boots':  'IKF 360 - Field of Dreams',
        'maidaan_ki_rani':  'IKF 360 - Field of Dreams',
        'ikf_trials':       'IKF Trials',
        'anayas_goal':      "Anaya's Goal",
    }

    # Custom icon methods
    def invictus_icon(self, obj):
        return self._bool_icon(obj.invictus_icare)

   
    def fc_icon(self, obj):
        return self._bool_icon(obj.future_champions)
    def bb_icon(self, obj):
        return self._bool_icon(obj.books_and_boots)
    def mr_icon(self, obj):
        return self._bool_icon(obj.maidaan_ki_rani)
    def trials_icon(self, obj):
        return self._bool_icon(obj.ikf_trials)
    def anaya_icon(self, obj):
        return self._bool_icon(obj.anayas_goal)

    invictus_icon.short_description = 'i-Care By Invictus'
    fc_icon.short_description = 'Field of Dreams (Future champion)'
    bb_icon.short_description = 'Field of Dreams (Books & Boots)'
    mr_icon.short_description = 'Field of Dreams (Maidaan Ki Raani)'
    trials_icon.short_description = 'IKF Trials'
    anaya_icon.short_description = "Anaya's Goal"

    def _bool_icon(self, value):
        if value:
            return format_html('<span style="color: green;">{}</span>', '✔️')
        return format_html('<span style="color: red;">{}</span>', '❌')

    # Other fields same as before
    list_filter = (
        NaariShaktiEditionFilter,   # Bihar / Rajasthan / India editions
        'status',
        'project',          # filter by admin-created DonateProject
        'created_at',
        'invictus_icare',
        'future_champions',
        'books_and_boots',
        'maidaan_ki_rani',
        'ikf_trials',
        'anayas_goal',
    )
    search_fields = ('name', 'email', 'mobile', 'razorpay_order_id', 'address', 'project__name')
    ordering = ('-created_at',)
    readonly_fields = ('created_at', 'updated_at')

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        formfield = super().formfield_for_dbfield(db_field, request, **kwargs)
        label = self.donation_purpose_labels.get(db_field.name)
        if label and formfield:
            formfield.label = label
        return formfield

    fieldsets = (
        ('Donor Info', {
            'fields': ('name', 'email', 'mobile', 'address', 'amount', 'photo')
        }),
        ('Donation Purpose', {
            'fields': (
                'invictus_icare',
                'future_champions',
                'books_and_boots',
                'maidaan_ki_rani',
                'ikf_trials',
                'anayas_goal',
            )
        }),
        ('Payment Info', {
            'fields': ('status', 'razorpay_order_id', 'razorpay_payment_id', 'razorpay_signature')
        }),
        ('Errors (if any)', {
            'classes': ('collapse',),
            'fields': (
                'error_code',
                'error_description',
                'error_source',
                'error_reason',
                'error_meta_order_id',
                'error_meta_payment_id'
            )
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at')
        }),
    )

    # Project-purpose boolean field names, in display order
    purpose_fields = (
        'invictus_icare',
        'future_champions',
        'books_and_boots',
        'maidaan_ki_rani',
        'ikf_trials',
        'anayas_goal',
    )

    # Maps legacy boolean field → DonateProject slug.
    # Look-up happens at request time so renaming the project in the
    # DonateProject admin instantly flows into the Donations list view.
    # If the slug of a project changes in admin, update it here too.
    purpose_to_slug = {
        'invictus_icare':   'donate-project-5',   # I-Care by Invictus
        'future_champions': 'donate-project-4',   # IKF 360 - Field of Dreams
        'books_and_boots':  'donate-project-4',   # IKF 360 - Field of Dreams
        'maidaan_ki_rani':  'donate-project-4',   # IKF 360 - Field of Dreams
        'ikf_trials':       'donate-project-1',   # IKF Trials
        'anayas_goal':      None,                  # no current project
    }

    def get_queryset(self, request):
        """Pre-build a {slug: name} cache once per request so the
        `donated_for` column doesn't issue an extra query per row."""
        qs = super().get_queryset(request)
        from .models import DonateProject
        slugs = {s for s in self.purpose_to_slug.values() if s}
        self._slug_to_name = {
            p.slug: p.name
            for p in DonateProject.objects.filter(slug__in=slugs).only('slug', 'name')
        }
        return qs

    def get_fieldsets(self, request, obj=None):
        """On the change form for an existing donation, only show the
        donation-purpose booleans that are True — so the admin sees just
        the project(s) this donation was made for, not all six unchecked
        siblings. New-donation form keeps the full set so the admin can
        still pick when creating manually."""
        fieldsets = super().get_fieldsets(request, obj)
        if obj is None:
            return fieldsets

        active = tuple(f for f in self.purpose_fields if getattr(obj, f, False))

        rebuilt = []
        for name, opts in fieldsets:
            if name == 'Donation Purpose':
                if not active:
                    # Nothing flagged True — skip the whole section
                    continue
                new_opts = dict(opts)
                new_opts['fields'] = active
                rebuilt.append((name, new_opts))
            else:
                rebuilt.append((name, opts))
        return tuple(rebuilt)




@admin.register(SupportIkfFeatureStrip)
class SupportIkfFeatureStrip(admin.ModelAdmin):
    list_display = ('id', 'keydata',  'heading', 'paragraph', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4',)
    search_fields = ('id', 'keydata', 'name')

@admin.register(PartnerSeasonSupportIKF)
class PartnerSeasonSupportIKF(admin.ModelAdmin):
    list_display = ('id', 'keydata', 'name', 'pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4')
    search_fields = ('id', 'keydata', 'name')

@admin.register(Calender)
class CalenderIKF(admin.ModelAdmin):
    list_display = ('id', 'date','second_date','third_date', 'project_id', 'project_city_id', 'project_city', 'project_state_id', 'project_state', 'project_country_id', 'project_country','pic', 'lang', 'attr1', 'attr2', 'attr3', 'attr4')
    list_filter = ('project','project_city','project_state')
    search_fields = ('id', 'project_city', 'project_state')

@admin.register(Finalists)
class FinalistsAdmin(admin.ModelAdmin):
    list_display = ('id', 'pic', 'name',)
    list_filter = ('name',)
    search_fields = ('name',)

@admin.register(Coaches)
class Coaches(admin.ModelAdmin):
    list_display = ('id', 'pic', 'name', 'description')
    list_filter = ('name',)
    search_fields = ('name',)

from django.contrib import admin
from .models import Scout

@admin.register(Scout)
class ScoutAdmin(admin.ModelAdmin):
    list_display = (
        'id',
        'ikf_id',
        'full_name',
        'location',
        'mobile_number',
        'gmail_id',
        'linkedin',
        'instagram',
        'facebook',
        'ikf_tshirt',
        'ishidden',
        'profile_photo'
    )
    search_fields = ('full_name', 'mobile_number', 'gmail_id')
    list_filter = ('ikf_tshirt', 'ishidden', 'location')


from .models import OfficialScoutingPartner

@admin.register(OfficialScoutingPartner)
class OfficialScoutingPartnerAdmin(admin.ModelAdmin):
    list_display = ['name', 'image']




from django.contrib import admin
from django.utils.html import format_html
from .models import IWLPartner

@admin.register(IWLPartner)
class IWLPartnerAdmin(admin.ModelAdmin):
    list_display = ('name', 'image_display')
    search_fields = ('name',)

    def image_display(self, obj):
        if obj.image:
            return format_html('<img src="{}" width="80" height="50" style="object-fit: contain;" />', obj.image.url)
        return "-"
    image_display.short_description = 'Image'


from django.contrib import admin
from .models import Faq

@admin.register(Faq)
class FaqAdmin(admin.ModelAdmin):
    list_display = ('page_name', 'priority', 'question')
    list_filter = ('page_name',)
    ordering = ('page_name', 'priority')



   



    



from .models import TermsSection

@admin.register(TermsSection)
class TermsSectionAdmin(admin.ModelAdmin):
    list_display        = ('order', 'title', 'is_active', 'updated_at')
    list_display_links  = ('title',)
    list_editable       = ('order', 'is_active')
    search_fields       = ('title', 'content')
    ordering            = ('order',)



from .models import RepOpenState, RepOpenCity


class RepOpenCityInline(admin.TabularInline):
    model               = RepOpenCity
    extra               = 1
    fields              = ('city', 'order', 'is_active')
    ordering            = ('order', 'city')
    verbose_name        = "City"
    verbose_name_plural = "Cities (add one row per city)"


@admin.register(RepOpenState)
class RepOpenStateAdmin(admin.ModelAdmin):
    list_display  = ('name', 'order', 'is_active', 'city_count')
    list_editable = ('order', 'is_active')
    search_fields = ('name',)
    ordering      = ('order', 'name')
    inlines       = [RepOpenCityInline]



from .models import IKFAchievement, ImpactReport

@admin.register(IKFAchievement)
class IKFAchievementAdmin(admin.ModelAdmin):
    list_display        = ('order', 'award_name', 'awarding_agency', 'year', 'is_active')
    list_display_links  = ('award_name',)
    list_editable       = ('order', 'is_active')
    search_fields       = ('award_name', 'awarding_agency', 'details')
    ordering            = ('order',)

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="height:50px;border-radius:6px;" />', obj.image.url)
        return "—"
    image_preview.short_description = "Preview"
    readonly_fields = ('image_preview',)
    fields = ('award_name', 'awarding_agency', 'year', 'details', 'image', 'image_preview', 'order', 'is_active')


@admin.register(ImpactReport)
class ImpactReportAdmin(admin.ModelAdmin):
    list_display       = ('order', 'title', 'year', 'show_in_side_view', 'is_active')
    list_display_links = ('title',)
    list_editable      = ('order', 'show_in_side_view', 'is_active')
    search_fields      = ('title', 'year')
    ordering           = ('order',)

    def cover_preview(self, obj):
        if obj.cover_image:
            return format_html('<img src="{}" style="height:60px;border-radius:6px;" />', obj.cover_image.url)
        return "—"
    cover_preview.short_description = "Cover"
    readonly_fields = ('cover_preview',)
    fields = ('title', 'year', 'cover_image', 'cover_preview', 'pdf_file', 'view_link', 'show_in_side_view', 'order', 'is_active')


from .models import OfficialCircularCard, OfficialCircularItem


class OfficialCircularItemAdminForm(forms.ModelForm):
    body = forms.CharField(
        required=False,
        widget=forms.Textarea(attrs={"rows": 12, "cols": 120}),
        help_text="Full description or body text",
    )

    class Meta:
        model = OfficialCircularItem
        fields = "__all__"

class OfficialCircularItemInline(admin.StackedInline):
    model   = OfficialCircularItem
    form    = OfficialCircularItemAdminForm
    extra   = 1
    fields  = ('title', 'tag', 'date', 'body', 'image', 'attachment', 'external_link', 'order', 'is_active')
    ordering = ('order', '-date')

@admin.register(OfficialCircularCard)
class OfficialCircularCardAdmin(admin.ModelAdmin):
    list_display       = ('order', 'title', 'slug', 'is_active')
    list_display_links = ('title',)
    list_editable      = ('order', 'is_active')
    search_fields      = ('title',)
    ordering           = ('order',)
    prepopulated_fields = {'slug': ('title',)}
    fields             = ('title', 'description', 'emoji', 'slug', 'order', 'is_active')
    inlines            = [OfficialCircularItemInline]

@admin.register(OfficialCircularItem)
class OfficialCircularItemAdmin(admin.ModelAdmin):
    form               = OfficialCircularItemAdminForm
    list_display       = ('order', 'card', 'title', 'tag', 'date', 'is_active')
    list_display_links = ('title',)
    list_editable      = ('order', 'is_active')
    list_filter        = ('card', 'is_active')
    search_fields      = ('title', 'body')
    ordering           = ('card', 'order', '-date')


from .models import IKFBroadcastItem

class IKFBroadcastBodyForm(forms.ModelForm):
    body = forms.CharField(
        required=False,
        widget=forms.Textarea(attrs={"rows": 10, "cols": 120}),
    )
    class Meta:
        model  = IKFBroadcastItem
        fields = "__all__"

@admin.register(IKFBroadcastItem)
class IKFBroadcastItemAdmin(admin.ModelAdmin):
    form           = IKFBroadcastBodyForm
    list_display   = ('__str__', 'date', 'is_active', 'created_at')
    list_editable  = ('is_active',)
    list_filter    = ('is_active',)
    search_fields  = ('heading', 'context', 'body')
    ordering       = ('-date', '-created_at')
    fields         = (
        'heading', 'context', 'date', 'image',
        'body',
        'cta_label', 'cta_url',
        'cta_label_2', 'cta_url_2',
        'is_active',
    )


from .models import HomeHeroSlide, HomeFootprintProject, HomeBackboneCard


@admin.register(HomeHeroSlide)
class HomeHeroSlideAdmin(admin.ModelAdmin):
    """
    Slides for the big rotating banner at the very top of the home page.
    Add as many slides as you like — they auto-rotate in order.

    Image sizes:
      • Background image: 1920×1080 px (16:9)
      • Character image:   800×1500 px portrait, transparent PNG
    """
    list_display  = ("order", "title_preview", "is_active", "character_position")
    list_editable = ("is_active",)
    list_filter   = ("is_active", "character_position")
    search_fields = ("title", "subtitle", "body", "tagline")
    ordering      = ("order", "id")
    fieldsets = (
        ("Order & visibility", {
            "fields": ("order", "is_active"),
        }),
        ("Text", {
            "fields": ("title", "subtitle", "body", "tagline"),
            "description": "What appears on the slide.",
        }),
        ("Images", {
            "fields": ("background_image", "character_image", "character_position"),
            "description": (
                "Recommended sizes — Background: 1920×1080 px (16:9). "
                "Character: 800×1500 px portrait, transparent PNG."
            ),
        }),
    )

    def title_preview(self, obj):
        return (obj.title[:70] + "…") if len(obj.title) > 70 else obj.title
    title_preview.short_description = "Title"


@admin.register(HomeFootprintProject)
class HomeFootprintProjectAdmin(admin.ModelAdmin):
    """
    Rows for the 'Footprint Projects' section under the hero carousel.
    Each row becomes one tab on the left + image panel on the right.
    Add unlimited rows.

    Image size: 1200×900 px (4:3). Image is cropped to fit, keep subject centred.
    """
    list_display  = ("order", "tab_title", "is_active", "explore_url")
    list_editable = ("is_active",)
    list_filter   = ("is_active",)
    search_fields = ("tab_title", "tab_subtitle", "pane_title", "pane_description")
    ordering      = ("order", "id")
    fieldsets = (
        ("Order & visibility", {
            "fields": ("order", "is_active"),
        }),
        ("Left tab (the small list on the left)", {
            "fields": ("tab_title", "tab_subtitle"),
        }),
        ("Right panel (image + headline + button)", {
            "fields": ("pane_title", "pane_description", "image", "explore_url"),
            "description": "Recommended image size: 1200×900 px (4:3). Pane title is optional — leave blank to reuse the tab title.",
        }),
    )


@admin.register(HomeBackboneCard)
class HomeBackboneCardAdmin(admin.ModelAdmin):
    """
    Cards in the 'Build the football ecosystem with IKF' section
    (IKF Backbone). Add unlimited cards. No image — text only.
    """
    list_display  = ("order", "title", "kicker", "is_active", "url")
    list_editable = ("is_active",)
    list_filter   = ("is_active",)
    search_fields = ("title", "kicker", "description")
    ordering      = ("order", "id")
    fieldsets = (
        ("Order & visibility", {
            "fields": ("order", "is_active"),
        }),
        ("Card content", {
            "fields": ("kicker", "title", "description", "arrow_label", "url"),
            "description": (
                "Kicker = the pill at the top of the card (e.g. 'Host City Trials'). "
                "Arrow label = the bottom action text (e.g. 'Become REP'). "
                "URL can be a path like '/rep/' or a full https://… link."
            ),
        }),
    )


from .models import PageStickyTitle


@admin.register(PageStickyTitle)
class PageStickyTitleAdmin(admin.ModelAdmin):
    """
    Simple page-title manager:
      1. Pick a page from the "Page" dropdown.
      2. Type the title you want in the sticky bar.
      3. Save. Uncheck "Is active" any time to revert that page to its default title.
    """
    list_display  = ("page_display", "title", "is_active", "updated_at")
    list_editable = ("is_active",)
    list_filter   = ("is_active",)
    search_fields = ("url_name", "title")
    fields        = ("url_name", "title", "is_active")

    def page_display(self, obj):
        return obj.get_url_name_display()
    page_display.short_description = "Page"
    page_display.admin_order_field = "url_name"


# ─────────────────────────────────────────────────────────────
# "Explore More" side-nav rail (admin-managed)
# ─────────────────────────────────────────────────────────────
from .models import ExploreNav, ExploreNavPage, ExploreNavLink
from .models import STICKY_PAGE_CHOICES, PAGE_CHOICES


def _explore_page_choices():
    """
    Build the 'Page' dropdown from EVERY real page on the site (all named URLs
    that don't need an argument), so nothing is missing. Friendly labels are
    used where we have one; otherwise the page's code is shown.
    """
    # Friendly label lookup, built from the curated choice lists.
    label_map = {}
    for key, lbl in list(STICKY_PAGE_CHOICES) + list(PAGE_CHOICES):
        label_map.setdefault(key, lbl)

    seen = set()
    choices = []
    try:
        from . import urls as main_urls
        for pat in getattr(main_urls, "urlpatterns", []):
            name = getattr(pat, "name", None)
            if not name or name in seen:
                continue
            # Skip pages that need a URL argument (slug/id/etc.).
            try:
                if pat.pattern.regex.groups:
                    continue
            except Exception:
                continue
            seen.add(name)
            label = label_map.get(name, name)
            choices.append((name, label if label == name else "%s  (%s)" % (label, name)))
    except Exception:
        pass

    # Safety net: include any curated pages we didn't discover above.
    for key, lbl in STICKY_PAGE_CHOICES:
        if key not in seen:
            seen.add(key)
            choices.append((key, "%s  (%s)" % (lbl, key)))

    choices.sort(key=lambda c: c[1].lower())
    return choices


class ExploreNavPageForm(forms.ModelForm):
    url_name = forms.ChoiceField(
        label="Page",
        help_text="Choose a page where this menu should appear.",
    )

    class Meta:
        model = ExploreNavPage
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["url_name"].choices = [("", "--- choose a page ---")] + _explore_page_choices()


class ExploreNavPageInline(admin.TabularInline):
    model = ExploreNavPage
    form = ExploreNavPageForm
    extra = 1
    verbose_name = "Page"
    verbose_name_plural = "1) Show this menu on these pages"


class ExploreNavLinkForm(forms.ModelForm):
    class Meta:
        model = ExploreNavLink
        fields = "__all__"
        labels = {
            "label":        "Menu text (what visitors see)",
            "external_url": "Link (paste the web address)",
            "group_header": "Group heading (optional)",
            "order":        "Position",
        }
        help_texts = {
            "label":        "The words shown in the menu, e.g. “Register for Trials”.",
            "external_url": "Paste the full web address, e.g. https://indiakhelofootball.com/season5 "
                            "(copy it from your browser’s address bar).",
            "group_header": "Optional. A small heading shown above this link to group links together, "
                            "e.g. “Players”. Leave blank if you don’t want a heading.",
            "order":        "Lower numbers show first (0, 1, 2 …).",
        }
        widgets = {
            "external_url": forms.TextInput(attrs={"size": 55, "placeholder": "https://…"}),
            "label":        forms.TextInput(attrs={"size": 40}),
            "group_header": forms.TextInput(attrs={"size": 30}),
        }


class ExploreNavLinkInline(admin.StackedInline):
    model = ExploreNavLink
    form = ExploreNavLinkForm
    extra = 2
    ordering = ("order", "id")
    # Only the plain-English fields. The developer-only "url_name"/"url_arg"
    # fields are intentionally hidden to keep this simple.
    fields = ("order", "group_header", "label", "external_url", "is_active")
    verbose_name = "Link"
    verbose_name_plural = "2) Links in the menu"


@admin.register(ExploreNav)
class ExploreNavAdmin(admin.ModelAdmin):
    """
    Build the 'Explore More' side menu without a developer:
      1. Name it (only you see this name).
      2. Type the tab text + heading visitors see.
      3. Choose the pages it appears on (or tick 'Show on all pages').
      4. Add the links — just a menu text and the web address for each.
    Any page without a menu keeps the built-in default menu.
    """
    list_display  = ("admin_name", "label", "heading", "scope", "is_active", "order", "updated_at")
    list_editable = ("is_active", "order")
    list_filter   = ("is_active", "apply_to_all_pages")
    search_fields = ("admin_name", "label", "heading", "links__label", "pages__url_name")
    inlines       = (ExploreNavPageInline, ExploreNavLinkInline)
    fieldsets = (
        ("Step 1 — Name this menu (only you see this)", {
            "fields": ("admin_name",),
            "description": "A name to help you find this menu later, e.g. “Explore menu for Player pages”. "
                           "Visitors never see this.",
        }),
        ("Step 2 — Where should it show?", {
            "fields": ("apply_to_all_pages",),
            "description": "Tick the box to show this menu on EVERY page except the Home page. "
                           "Leave it un-ticked to choose specific pages in the “Show this menu on these pages” "
                           "section lower down.",
        }),
        ("Step 3 — What visitors see", {
            "fields": ("label", "heading", "show_home_link"),
            "description": "“Label” is the small vertical tab on the right edge of the page. "
                           "“Heading” is the title at the top of the menu once it opens. "
                           "Both can just be “Explore More”.",
        }),
        ("Settings", {
            "fields": ("order", "is_active"),
            "classes": ("collapse",),
            "description": "Leave these as they are unless you know you need them. "
                           "Un-tick “Is active” to temporarily switch this menu off.",
        }),
    )

    def scope(self, obj):
        if obj.apply_to_all_pages:
            return "All pages (except Home)"
        names = [p.get_url_name_display() for p in obj.pages.all()]
        return ", ".join(names) if names else "— (no pages)"
    scope.short_description = "Shows on"

    def page_list(self, obj):
        names = [p.get_url_name_display() for p in obj.pages.all()]
        return ", ".join(names) if names else "—"
    page_list.short_description = "Pages"


# ─────────────────────────────────────────────────────────────
# IKF Career 360 — The Parent Conversation (quiz responses)
# ─────────────────────────────────────────────────────────────
import csv as _csv
import io as _io
from django.urls import path as _path
from django.http import HttpResponse as _HttpResponse
from .models import ParentConversation


PARENT_CONVO_QUESTIONS = {
    "q1": "Q1. How did your child first fall in love with football?",
    "q2": "Q2. When your child talks about football, what do you hear most?",
    "q3": "Q3. Your child is 22. Football has given them everything it can. What would make you feel the journey was worth it?",
    "q4": "Q4. On a normal week, what does your support look like?",
    "q5": "Q5. What is the hardest part of supporting your child's football journey right now?",
    "q6": "Q6. Of every 100 children who play football seriously in India — how many do you think turn professional?",
    "q7": "Q7. If your child gives ten years to this game and does not play professionally — what would you want football to have given them?",
    "q8": "Q8. IKF Career 360 exists because we made a decision early — we would not build a programme that only serves the 1%. Every child in our ecosystem has a pathway. On the field and beyond it. Does that matter to you?",
}

PARENT_CONVO_OPTIONS = {
    "q1": {
        "A": "My child just started playing one day and never stopped.",
        "B": "We introduced them to it — and they took it from there.",
        "C": "A coach or school spotted something in them.",
        "D_open": "It's a longer story (see open answer)",
    },
    "q2": {
        "A": "I want to play professionally. This is my life.",
        "B": "I love football. I don't know where it leads yet.",
        "C": "I play because it makes me happy — nothing more.",
        "D": "Honestly, I'm not sure. They haven't said.",
    },
    "q3": {
        "A": "They're playing at a professional or semi-professional level.",
        "B": "They're earning a living inside football — coaching, management, media.",
        "C": "They grew into a confident, disciplined, resilient person — football did that.",
        "D": "All three. I won't settle for less.",
    },
    "q4": {
        "A": "I'm there — planning, tracking, making sure everything lines up.",
        "B": "I provide what's needed — fees, travel, kit — but day-to-day is their responsibility.",
        "C": "I show up for matches and big moments. The rest is up to them.",
        "D": "Honestly, I'm still figuring out what support should look like.",
    },
    "q6": {
        "A": "Fewer than 1",
        "B": "Around 5",
        "C": "Around 10",
        "D": "More than 10",
    },
    "q7": {
        "A": "A career inside football — coaching, scouting, sports management, media.",
        "B": "The values. Discipline, teamwork, resilience. Those will serve them anywhere.",
        "C": "Both. A connection to the game and the character it builds.",
        "D": "Honestly — if they don't play professionally, I'll need time to think about what comes next.",
    },
    "q8": {
        "A": "Yes — that's exactly the kind of programme I want my child in.",
        "B": "I hadn't thought about it this way — but yes, it matters.",
        "C": "My focus right now is on the professional pathway. We can think about the rest later.",
    },
}


def _option_text(qkey, code):
    if not code:
        return "—"
    return PARENT_CONVO_OPTIONS.get(qkey, {}).get(code, code)


@admin.register(ParentConversation)
class ParentConversationAdmin(admin.ModelAdmin):
    change_list_template = "admin/main/parentconversation/change_list.html"

    list_display = (
        "created_at",
        "parent_name",
        "parent_email",
        "child_age",
        "child_gender",
        "profile_badge",
    )
    list_filter   = ("profile", "child_gender", "created_at")
    search_fields = ("parent_name", "parent_email")

    readonly_fields = (
        "created_at",
        "parent_name", "parent_email", "child_age", "child_gender",
        "q1_full", "q1_origin_open",
        "q2_full",
        "q3_full",
        "q4_full",
        "q5_full",
        "q6_full",
        "q7_full",
        "q8_full",
        "profile",
    )

    fieldsets = (
        ("Parent intake", {
            "fields": ("created_at", "parent_name", "parent_email", "child_age", "child_gender"),
        }),
        ("Movement 1 — Your child's story", {
            "fields": (
                "q1_full",
                "q1_origin_open",
                "q2_full",
            ),
        }),
        ("Movement 2 — Your story as a parent", {
            "fields": ("q3_full", "q4_full", "q5_full"),
        }),
        ("Movement 3 — The honest conversation", {
            "fields": ("q6_full", "q7_full", "q8_full"),
        }),
        ("Result", {
            "fields": ("profile",),
        }),
    )

    actions = ["export_selected_csv", "export_selected_xlsx"]

    def has_add_permission(self, request):
        return False

    # ── Full-text question/answer displays ────────────────────
    @admin.display(description=PARENT_CONVO_QUESTIONS["q1"])
    def q1_full(self, obj):
        return _option_text("q1", obj.q1_origin)

    @admin.display(description=PARENT_CONVO_QUESTIONS["q2"])
    def q2_full(self, obj):
        return _option_text("q2", obj.q2_child_voice)

    @admin.display(description=PARENT_CONVO_QUESTIONS["q3"])
    def q3_full(self, obj):
        return _option_text("q3", obj.q3_worth_it)

    @admin.display(description=PARENT_CONVO_QUESTIONS["q4"])
    def q4_full(self, obj):
        return _option_text("q4", obj.q4_support)

    @admin.display(description=PARENT_CONVO_QUESTIONS["q5"])
    def q5_full(self, obj):
        return obj.q5_hardest_part or "—"

    @admin.display(description=PARENT_CONVO_QUESTIONS["q6"])
    def q6_full(self, obj):
        return _option_text("q6", obj.q6_pro_estimate)

    @admin.display(description=PARENT_CONVO_QUESTIONS["q7"])
    def q7_full(self, obj):
        return _option_text("q7", obj.q7_ten_years)

    @admin.display(description=PARENT_CONVO_QUESTIONS["q8"])
    def q8_full(self, obj):
        return _option_text("q8", obj.q8_ecosystem)

    @admin.display(description="Profile", ordering="profile")
    def profile_badge(self, obj):
        colors = {
            "aligned":   ("#085041", "#e1f5ee"),
            "potential": ("#8c5a00", "#fff6d6"),
            "atrisk":    ("#c64a1d", "#faece7"),
        }
        fg, bg = colors.get(obj.profile, ("#1e293b", "#f1f5f9"))
        return format_html(
            '<span style="display:inline-block; padding:3px 10px; border-radius:999px; '
            'background:{}; color:{}; font-weight:700; font-size:11px; letter-spacing:.06em; '
            'text-transform:uppercase;">{}</span>',
            bg, fg, obj.get_profile_display(),
        )

    # ── CSV / XLSX export ─────────────────────────────────────
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            _path(
                "export-csv/",
                self.admin_site.admin_view(self.export_all_csv),
                name="main_parentconversation_export_csv",
            ),
            _path(
                "export-xlsx/",
                self.admin_site.admin_view(self.export_all_xlsx),
                name="main_parentconversation_export_xlsx",
            ),
        ]
        return custom + urls

    @staticmethod
    def _export_headers():
        return [
            "Submitted at",
            "Parent name",
            "Parent email",
            "Child age",
            "Child gender",
            "RESULT (Profile)",
            "Q1 answer",
            "Q1 longer story",
            "Q2 answer",
            "Q3 answer",
            "Q4 answer",
            "Q5 hardest part (open)",
            "Q6 answer",
            "Q7 answer",
            "Q8 answer",
        ]

    @staticmethod
    def _export_row(r):
        return [
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            r.parent_name,
            r.parent_email,
            r.child_age,
            r.get_child_gender_display(),
            r.get_profile_display(),
            _option_text("q1", r.q1_origin),
            r.q1_origin_open or "(not applicable — option not selected)",
            _option_text("q2", r.q2_child_voice),
            _option_text("q3", r.q3_worth_it),
            _option_text("q4", r.q4_support),
            r.q5_hardest_part or "",
            _option_text("q6", r.q6_pro_estimate),
            _option_text("q7", r.q7_ten_years),
            _option_text("q8", r.q8_ecosystem),
        ]

    def _write_csv(self, queryset):
        response = _HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            'attachment; filename="parent_conversation_responses.csv"'
        )
        writer = _csv.writer(response)

        writer.writerow(self._export_headers())
        for r in queryset.order_by("-created_at"):
            writer.writerow(self._export_row(r))

        # Question key reference, written below the data so the data table
        # columns remain readable in Excel / Numbers / Sheets.
        writer.writerow([])
        writer.writerow(["QUESTION KEY (full question text)"])
        for qkey in ("q1", "q2", "q3", "q4", "q5", "q6", "q7", "q8"):
            writer.writerow([qkey.upper(), PARENT_CONVO_QUESTIONS[qkey]])

        return response

    def _write_xlsx(self, queryset):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Parent Conversation"

        headers = self._export_headers()
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor="FF1F2937")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for r in queryset.order_by("-created_at"):
            ws.append(self._export_row(r))

        # Auto-size columns based on content (capped for readability)
        for col_idx, _header in enumerate(headers, start=1):
            max_len = len(str(_header))
            for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx,
                                          min_row=2, values_only=True):
                value = row_cells[0]
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 12), 50)

        ws.freeze_panes = "A2"

        # Question key reference sheet
        key_ws = wb.create_sheet("Question Key")
        key_ws.append(["Question", "Full text"])
        key_ws.cell(row=1, column=1).font = header_font
        key_ws.cell(row=1, column=1).fill = header_fill
        key_ws.cell(row=1, column=2).font = header_font
        key_ws.cell(row=1, column=2).fill = header_fill
        for qkey in ("q1", "q2", "q3", "q4", "q5", "q6", "q7", "q8"):
            key_ws.append([qkey.upper(), PARENT_CONVO_QUESTIONS[qkey]])
        key_ws.column_dimensions["A"].width = 10
        key_ws.column_dimensions["B"].width = 90

        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = _HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            'attachment; filename="parent_conversation_responses.xlsx"'
        )
        return response

    def export_all_csv(self, request):
        return self._write_csv(ParentConversation.objects.all())

    def export_all_xlsx(self, request):
        return self._write_xlsx(ParentConversation.objects.all())

    @admin.action(description="Download selected responses as CSV")
    def export_selected_csv(self, request, queryset):
        return self._write_csv(queryset)

    @admin.action(description="Download selected responses as Excel (.xlsx)")
    def export_selected_xlsx(self, request, queryset):
        return self._write_xlsx(queryset)


from .models import OfflineDonationClaim


@admin.register(OfflineDonationClaim)
class OfflineDonationClaimAdmin(admin.ModelAdmin):
    list_display  = ('created_at', 'project', 'method', 'donor_display', 'amount', 'status')
    list_filter   = ('status', 'method', 'project', 'anonymous', 'created_at')
    search_fields = ('name', 'email', 'mobile', 'organisation', 'project', 'amount')
    list_editable = ('status',)
    date_hierarchy = 'created_at'
    readonly_fields = ('created_at', 'updated_at')
    fieldsets = (
        ('Claim', {
            'fields': ('project', 'method', 'amount', 'status', 'admin_notes')
        }),
        ('Donor', {
            'fields': ('anonymous', 'name', 'organisation', 'email', 'mobile')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at')
        }),
    )

    def donor_display(self, obj):
        if obj.anonymous:
            return 'Anonymous'
        return obj.name or obj.email or obj.mobile or '-'
    donor_display.short_description = 'Donor'


from .models import DonateProject, DonateProjectSection


class DonateProjectSectionInline(admin.StackedInline):
    model = DonateProjectSection
    extra = 0
    fields = ('display_order', 'heading', 'content', 'image')
    ordering = ('display_order', 'id')


@admin.register(DonateProject)
class DonateProjectAdmin(admin.ModelAdmin):
    list_display  = ('display_order', 'name', 'slug', 'is_active', 'razorpay_enabled', 'upi_enabled', 'bank_enabled', 'preset_amounts', 'updated_at')
    list_editable = ('name', 'is_active', 'razorpay_enabled', 'upi_enabled', 'bank_enabled', 'preset_amounts')
    list_display_links = ('display_order',)
    list_filter   = ('is_active', 'razorpay_enabled', 'upi_enabled', 'bank_enabled')
    search_fields = ('name', 'description', 'slug')
    ordering      = ('display_order', 'id')
    prepopulated_fields = {'slug': ('name',)}
    inlines = [DonateProjectSectionInline]
    fieldsets = (
        ('Card content', {
            'fields': ('name', 'slug', 'description', 'read_more_url'),
            'description': "read_more_url is used as the 'Know more' link ONLY when the detail page below has no content (no banner, no intro, no sections)."
        }),
        ('Detail page content', {
            'fields': ('detail_banner', 'detail_intro'),
            'description': "Top of the project detail page. The 'Sections' below stack under this banner area."
        }),
        ('Display', {
            'fields': ('display_order', 'is_active')
        }),
        ('Donor panel options', {
            'fields': ('preset_amounts', 'razorpay_enabled', 'upi_enabled', 'bank_enabled')
        }),
    )


from .models import DonateSettings


@admin.register(DonateSettings)
class DonateSettingsAdmin(admin.ModelAdmin):
    fieldsets = (
        ('Trust markers', {
            'fields': ('reg_80g', 'csr_number')
        }),
        ('UPI', {
            'fields': ('upi_id', 'upi_qr_image')
        }),
        ('Bank Transfer', {
            'fields': ('bank_account_name', 'bank_account_number', 'bank_ifsc', 'bank_swift', 'bank_name', 'bank_branch')
        }),
        ('Contact', {
            'fields': ('whatsapp_number', 'donate_email')
        }),
    )

    def has_add_permission(self, request):
        return not DonateSettings.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        DonateSettings.load()
        return super().changelist_view(request, extra_context=extra_context)
